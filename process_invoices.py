# -*- coding: utf-8 -*-
"""
process_invoices.py
Wersja z integracją Google Document AI (Invoice Parser).
- Jeżeli są ustawione zmienne/sekrety Google -> czyta prawdziwe faktury.
- Jeżeli nie -> działa na stubach (placeholder), żeby pipeline był stabilny.
- Walidacja sum, zapis do Excela (3 arkusze), tryb DRY_RUN.

Uruchomienia:
  python process_invoices.py --once
  DRY_RUN=true python process_invoices.py --once
  python process_invoices.py               # pętla co 5 min
"""

from __future__ import annotations

import os
import sys
import json
import time
import math
import logging
import argparse
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional

# =============== KONFIG ===============

INBOX_DIR = Path(os.getenv("INBOX_DIR", "Skopiowane_faktury")).resolve()
OUT_XLSX = Path(os.getenv("OUT_XLSX", "Szablon_Faktury_AI_v4.xlsx")).resolve()
LOOP_SLEEP_SEC = int(os.getenv("LOOP_SLEEP_SEC", "300"))  # 5 min
DRY_RUN = os.getenv("DRY_RUN", "").lower() in {"1", "true", "yes"}
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()

# Google Document AI (opcjonalnie, włącz jeśli masz sekrety/zależności)
DOC_AI_PROJECT = os.getenv("DOC_AI_PROJECT")          # np. "my-project"
DOC_AI_LOCATION = os.getenv("DOC_AI_LOCATION", "eu")  # np. "eu"
DOC_AI_PROCESSOR = os.getenv("DOC_AI_PROCESSOR")      # np. "1234567890abcdef"
GOOGLE_APPLICATION_CREDENTIALS_JSON = os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON")

SUPPORTED_EXT = {".pdf", ".jpg", ".jpeg", ".png"}

# =============== LOGGING ===============

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("process_invoices")

# =============== MODELE DANYCH ===============

@dataclass
class LineItem:
    description: str
    quantity: float
    unit_price: float
    net: float
    vat_rate: float
    vat: float
    gross: float


@dataclass
class Invoice:
    number: str
    issue_date: str  # ISO yyyy-mm-dd
    seller: str
    buyer: str
    currency: str
    total_net: float
    total_vat: float
    total_gross: float
    line_items: List[LineItem]

    def as_header_row(self) -> Dict[str, Any]:
        return {
            "number": self.number,
            "issue_date": self.issue_date,
            "seller": self.seller,
            "buyer": self.buyer,
            "currency": self.currency,
            "total_net": self.total_net,
            "total_vat": self.total_vat,
            "total_gross": self.total_gross,
        }

# =============== WALIDACJA SUM ===============

def validate_totals(data: Dict[str, Any], tol: float = 0.01) -> Tuple[bool, List[str]]:
    total_net = float(data.get("total_net", 0) or 0)
    total_vat = float(data.get("total_vat", 0) or 0)
    total_gross = float(data.get("total_gross", 0) or 0)

    items = data.get("line_items", []) or []
    net_sum = sum(float(i.get("net", 0) or 0) for i in items)
    vat_sum = sum(float(i.get("vat", 0) or 0) for i in items)
    gross_sum = sum(float(i.get("gross", 0) or 0) for i in items)

    errors: List[str] = []
    if abs(net_sum - total_net) > tol:
        errors.append(f"Mismatch net: {net_sum:.2f} vs {total_net:.2f}")
    if abs(vat_sum - total_vat) > tol:
        errors.append(f"Mismatch vat: {vat_sum:.2f} vs {total_vat:.2f}")
    if abs(gross_sum - total_gross) > tol:
        errors.append(f"Mismatch gross: {gross_sum:.2f} vs {total_gross:.2f}")

    return len(errors) == 0, errors

# =============== DOKUMENTY – PLIKI ===============

def find_new_files(folder: Path) -> List[Path]:
    folder.mkdir(parents=True, exist_ok=True)
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_EXT]
    return sorted(files)

def read_file_bytes(p: Path) -> bytes:
    return p.read_bytes()

# =============== GOOGLE DOC AI – KLIENT (opcjonalnie) ===============

def _have_docai() -> bool:
    return bool(DOC_AI_PROJECT and DOC_AI_LOCATION and DOC_AI_PROCESSOR and GOOGLE_APPLICATION_CREDENTIALS_JSON)

def _build_docai_client():
    """
    Tworzy klienta Document AI z JSON-a trzymanego w sekrecie
    GOOGLE_APPLICATION_CREDENTIALS_JSON (bez zapisywania pliku na dysk).
    """
    from google.oauth2 import service_account
    from google.api_core.client_options import ClientOptions
    from google.cloud import documentai

    info = json.loads(GOOGLE_APPLICATION_CREDENTIALS_JSON)
    # Uwaga: jeżeli chcesz rozliczać w konkretnym projekcie rozliczeniowym, możesz
    # wstawić "quota_project_id" = DOC_AI_PROJECT (o ile to poprawny billing project).
    if "quota_project_id" not in info and DOC_AI_PROJECT:
        info["quota_project_id"] = DOC_AI_PROJECT

    creds = service_account.Credentials.from_service_account_info(
        info,
        scopes=[
            "https://www.googleapis.com/auth/cloud-platform",
        ],
    )

    endpoint = f"{DOC_AI_LOCATION}-documentai.googleapis.com"
    client = documentai.DocumentProcessorServiceClient(
        client_options=ClientOptions(api_endpoint=endpoint),
        credentials=creds,
    )
    return client

# =============== PARSING FAKTURY: DOC AI lub STUB ===============

def parse_invoice_docai(file_path: Path) -> Optional[Invoice]:
    """
    Parsuje fakturę przy pomocy Document AI (Invoice Parser).
    Zwraca Invoice albo None (w razie błędu).
    """
    try:
        from google.cloud import documentai
    except Exception as e:
        log.warning("Brak pakietu google-cloud-documentai: %s", e)
        return None

    try:
        client = _build_docai_client()
        name = client.processor_path(DOC_AI_PROJECT, DOC_AI_LOCATION, DOC_AI_PROCESSOR)

        raw = read_file_bytes(file_path)
        # Uwaga: Document AI łyka PDF, TIFF, oraz pojedyncze obrazy (JPEG/PNG).
        # Dla wielostronicowych JPG/PNG zaleca się konwersję do PDF – tu zakładamy
        # że jest pojedynczy obraz albo PDF.
        request = {
            "name": name,
            "raw_document": {
                "content": raw,
                "mime_type": _guess_mime(file_path.suffix.lower()),
            },
        }

        result = client.process_document(request=request)
        doc = result.document

        # Proste wyciąganie danych
        number = _docai_find_first_text(doc, ["invoice_id", "invoice_number"]) or f"INV-{file_path.stem}"
        issue_date = (_docai_find_first_text(doc, ["invoice_date", "issue_date"]) or
                      datetime.now().date().isoformat())
        seller = _docai_find_first_text(doc, ["supplier_name", "seller", "supplier"])
        buyer = _docai_find_first_text(doc, ["customer_name", "buyer", "customer"])
        currency = _docai_find_first_text(doc, ["currency"]) or "PLN"

        # Suma
        total_net = _docai_find_first_money(doc, ["net_amount", "subtotal_amount"]) or 0.0
        total_vat = _docai_find_first_money(doc, ["total_tax_amount", "tax_amount"]) or 0.0
        total_gross = _docai_find_first_money(doc, ["total_amount"]) or (
            total_net + total_vat if (total_net or total_vat) else 0.0
        )

        # Pozycje – Document AI ma tabelę line_items
        items = _docai_parse_line_items(doc, currency_hint=currency)
        if not items:
            # awaryjnie – choćby 1 pozycja z kwotami
            items = [LineItem(description="(no-items)", quantity=1.0, unit_price=total_gross,
                              net=total_net or total_gross, vat_rate=_guess_vat_rate(total_net, total_vat),
                              vat=total_vat, gross=total_gross)]

        inv = Invoice(
            number=number,
            issue_date=issue_date,
            seller=seller or "",
            buyer=buyer or "",
            currency=currency or "PLN",
            total_net=round(float(total_net), 2),
            total_vat=round(float(total_vat), 2),
            total_gross=round(float(total_gross), 2),
            line_items=items,
        )
        return inv
    except Exception as e:
        log.exception("DocAI parsing failed for %s: %s", file_path.name, e)
        return None

def parse_invoice_stub(file_path: Path) -> Invoice:
    """Fallback – sztuczne dane, żeby przepływ był stabilny."""
    items = [
        LineItem(description=f"Pozycja 1 z {file_path.name}", quantity=1.0, unit_price=100.0,
                 net=100.0, vat_rate=23.0, vat=23.0, gross=123.0),
        LineItem(description=f"Pozycja 2 z {file_path.name}", quantity=2.0, unit_price=50.0,
                 net=100.0, vat_rate=23.0, vat=23.0, gross=123.0),
    ]
    total_net = sum(i.net for i in items)
    total_vat = sum(i.vat for i in items)
    total_gross = sum(i.gross for i in items)

    return Invoice(
        number=f"INV-{file_path.stem}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        issue_date=datetime.now().date().isoformat(),
        seller="Acme Sp. z o.o.",
        buyer="Twoja Firma Sp. z o.o.",
        currency="PLN",
        total_net=total_net,
        total_vat=total_vat,
        total_gross=total_gross,
        line_items=items,
    )

def _guess_mime(ext: str) -> str:
    if ext == ".pdf": return "application/pdf"
    if ext in {".jpg", ".jpeg"}: return "image/jpeg"
    if ext == ".png": return "image/png"
    return "application/octet-stream"

def _docai_find_first_text(doc, field_names: List[str]) -> Optional[str]:
    # Entity extraction – Document AI klasyfikuje pola jako entities
    for ent in getattr(doc, "entities", []) or []:
        n = (ent.type_ or "").lower()
        if n in (fn.lower() for fn in field_names):
            val = (ent.mention_text or ent.normalized_value.text or "").strip()
            if val:
                return val
    return None

def _docai_find_first_money(doc, field_names: List[str]) -> Optional[float]:
    for ent in getattr(doc, "entities", []) or []:
        n = (ent.type_ or "").lower()
        if n in (fn.lower() for fn in field_names):
            # money_value or normalized_value.{money_value}
            try:
                if ent.normalized_value and ent.normalized_value.money_value:
                    mv = ent.normalized_value.money_value
                    amount = (mv.units or 0) + (mv.nanos or 0) / 1e9
                    return float(amount)
            except Exception:
                pass
            # fallback – parsuj tekst
            txt = (ent.mention_text or "").replace(",", ".")
            try:
                return float("".join(ch for ch in txt if (ch.isdigit() or ch in ".-")))
            except Exception:
                continue
    return None

def _docai_parse_line_items(doc, currency_hint: str = "PLN") -> List[LineItem]:
    """Próba wyciągnięcia pozycji z entities/tables. Minimalny, bezpieczny parser."""
    out: List[LineItem] = []
    # 1) Spróbuj line_item entities:
    for ent in getattr(doc, "entities", []) or []:
        if (ent.type_ or "").lower() in {"line_item", "lineItem", "lineitem"}:
            desc = _entity_child_text(ent, {"description", "item_description"}) or (ent.mention_text or "").strip()
            qty = _entity_child_float(ent, {"quantity", "qty"}) or 1.0
            unit_price = _entity_child_money(ent, {"unit_price", "unit_price_amount"}) or 0.0
            net = _entity_child_money(ent, {"net_amount", "subtotal_amount"}) or (qty * unit_price)
            vat = _entity_child_money(ent, {"tax_amount"}) or 0.0
            gross = _entity_child_money(ent, {"amount", "total_amount"}) or (net + vat)
            vat_rate = _guess_vat_rate(net, vat)
            out.append(LineItem(
                description=desc or "(line item)",
                quantity=float(qty),
                unit_price=float(unit_price),
                net=float(net),
                vat_rate=float(vat_rate),
                vat=float(vat),
                gross=float(gross),
            ))

    if out:
        return out

    # 2) Awaryjnie: spróbuj z tabel
    # (pełny parser tabel to kilka-kilkanaście ekranów – tutaj minimalny fallback)
    for page in getattr(doc, "pages", []) or []:
        for table in getattr(page, "tables", []) or []:
            try:
                headers = ["".join(_layout_text(doc, cell.layout) for cell in row.cells).lower()
                           for row in table.header_rows]
                # poszukaj pierwszego wiersza danych
                for row in table.body_rows:
                    texts = ["".join(_layout_text(doc, cell.layout) for cell in row.cells) for cell in row.cells]
                    if not texts:
                        continue
                    # heurystyka: opis + kwoty
                    desc = (texts[0] or "").strip() or "(line item)"
                    nums = [_safe_money(t) for t in texts[1:]]
                    if all(v is None for v in nums):
                        continue
                    # spróbuj rozpoznać kolumny
                    qty = nums[0] or 1.0
                    unit_price = (nums[1] if len(nums) > 1 else None) or 0.0
                    net = (nums[2] if len(nums) > 2 else None) or (qty * unit_price)
                    vat = (nums[3] if len(nums) > 3 else 0.0)
                    gross = (nums[4] if len(nums) > 4 else (net + vat))
                    out.append(LineItem(
                        description=desc, quantity=qty, unit_price=unit_price,
                        net=net, vat_rate=_guess_vat_rate(net, vat), vat=vat, gross=gross
                    ))
            except Exception:
                continue
    return out

def _entity_child_text(ent, names: set) -> Optional[str]:
    for prop in getattr(ent, "properties", []) or []:
        if (prop.type_ or "").lower() in {n.lower() for n in names}:
            txt = (prop.mention_text or prop.normalized_value.text or "").strip()
            if txt:
                return txt
    return None

def _entity_child_float(ent, names: set) -> Optional[float]:
    t = _entity_child_text(ent, names)
    if t is None:
        return None
    t = t.replace(",", ".")
    try:
        return float("".join(ch for ch in t if (ch.isdigit() or ch in ".-")))
    except Exception:
        return None

def _entity_child_money(ent, names: set) -> Optional[float]:
    # money z normalized_value ma units/nanos
    for prop in getattr(ent, "properties", []) or []:
        if (prop.type_ or "").lower() in {n.lower() for n in names}:
            try:
                if prop.normalized_value and prop.normalized_value.money_value:
                    mv = prop.normalized_value.money_value
                    return float((mv.units or 0) + (mv.nanos or 0) / 1e9)
            except Exception:
                pass
            txt = (prop.mention_text or "").replace(",", ".")
            try:
                return float("".join(ch for ch in txt if (ch.isdigit() or ch in ".-")))
            except Exception:
                continue
    return None

def _guess_vat_rate(net: float, vat: float) -> float:
    try:
        if net and vat is not None:
            r = (vat / net) * 100.0
            # zaokrąglij do typowych stawek w PL
            for cand in (23.0, 8.0, 5.0, 0.0):
                if abs(r - cand) < 1.0:
                    return cand
            return round(r, 2)
    except Exception:
        pass
    return 0.0

def _layout_text(doc, layout) -> str:
    # łączy tokeny wg segmentów – minimalny ekstraktor
    out = []
    for seg in getattr(layout, "text_anchor", {}).get("text_segments", []) or []:
        start = int(seg.start_index or 0)
        end = int(seg.end_index or 0)
        out.append((doc.text or "")[start:end])
    return "".join(out)

def _safe_money(s: str) -> Optional[float]:
    try:
        s = (s or "").replace(",", ".")
        return float("".join(ch for ch in s if (ch.isdigit() or ch in ".-")))
    except Exception:
        return None

# =============== ZAPIS DO EXCEL ===============

def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as e:
        raise RuntimeError("Brak pakietu 'openpyxl'. Zainstaluj: pip install openpyxl") from e

def write_to_excel(invoices: List[Invoice], out_xlsx: Path) -> None:
    ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Dane"
    headers1 = ["number", "issue_date", "seller", "buyer", "currency", "total_net", "total_vat", "total_gross"]
    ws1.append(headers1)
    for inv in invoices:
        ws1.append([inv.number, inv.issue_date, inv.seller, inv.buyer, inv.currency,
                    inv.total_net, inv.total_vat, inv.total_gross])

    ws2 = wb.create_sheet("Pozycje")
    headers2 = ["invoice_number", "description", "quantity", "unit_price", "net", "vat_rate", "vat", "gross"]
    ws2.append(headers2)
    for inv in invoices:
        for li in inv.line_items:
            ws2.append([inv.number, li.description, li.quantity, li.unit_price, li.net, li.vat_rate, li.vat, li.gross])

    ws3 = wb.create_sheet("Koszty_surowcow")
    ws3.append(["invoice_number", "category", "amount", "note"])

    for ws in (ws1, ws2, ws3):
        for col_idx, _ in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_xlsx))
    log.info("Zapisano Excel: %s", out_xlsx)

# =============== GŁÓWNY PRZEPŁYW ===============

def process_one_file(file_path: Path) -> Optional[Invoice]:
    """
    Jeśli jest Document AI -> użyj go; inaczej fallback do stubów.
    """
    try:
        if _have_docai():
            inv = parse_invoice_docai(file_path)
            if inv:
                log.info("DocAI OK: %s → %s, %s PLN", file_path.name, inv.number, inv.total_gross)
                return inv
            log.warning("DocAI zwrócił None, używam stubu dla %s", file_path.name)
        inv = parse_invoice_stub(file_path)
        return inv
    except Exception as e:
        log.exception("Błąd podczas przetwarzania %s: %s", file_path, e)
        return None

def run_once() -> None:
    files = find_new_files(INBOX_DIR)
    if not files:
        log.info("Brak plików w %s", INBOX_DIR)
        return

    invoices: List[Invoice] = []
    for f in files:
        inv = process_one_file(f)
        if inv:
            ok, errs = validate_totals({
                "total_net": inv.total_net,
                "total_vat": inv.total_vat,
                "total_gross": inv.total_gross,
                "line_items": [asdict(li) for li in inv.line_items],
            })
            if not ok:
                log.warning("Walidacja NIE przeszła dla %s: %s", f.name, errs)
            invoices.append(inv)

    if not invoices:
        log.info("Brak danych do zapisu.")
        return

    if DRY_RUN:
        log.info("[DRY_RUN] Podgląd danych – zapis XLSX pominięty.")
        for inv in invoices:
            log.info("HEADER: %s", json.dumps(inv.as_header_row(), ensure_ascii=False))
            for li in inv.line_items:
                log.info("LINE: %s", json.dumps(asdict(li), ensure_ascii=False))
        return

    write_to_excel(invoices, OUT_XLSX)

def run_loop() -> None:
    log.info("Start pętli – co %d s", LOOP_SLEEP_SEC)
    while True:
        run_once()
        time.sleep(LOOP_SLEEP_SEC)

# =============== CLI ===============

def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Przetwarzanie faktur (DocAI-ready, DRY_RUN, Excel).")
    p.add_argument("--once", action="store_true", help="Wykonaj jeden przebieg i zakończ.")
    return p.parse_args(argv)

def main(argv: List[str]) -> int:
    args = parse_args(argv)
    log.info("Konfig: INBOX_DIR=%s, OUT_XLSX=%s, DRY_RUN=%s, DOC_AI=%s",
             INBOX_DIR, OUT_XLSX, DRY_RUN, _have_docai())
    INBOX_DIR.mkdir(parents=True, exist_ok=True)
    try:
        if args.once:
            run_once()
        else:
            run_loop()
        return 0
    except KeyboardInterrupt:
        log.info("Przerwano przez użytkownika.")
        return 0
    except Exception as e:
        log.exception("Błąd krytyczny: %s", e)
        return 1

if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
